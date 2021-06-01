# @Time:2021/5/7 3:22 下午
# @Author:liuyue
# @Email:liuyue@soyoung.com
# @coding:utf-8

"""
excel

openpyxl：.xlsx的读写操作

测试数据：事先写入数据到excel
平时操作excel的流程(3个对象)：
    工作簿(Workbook)
    表单(Sheet)
    单元格(Cell)

1、准备测试数据
2、load_workbook模块，去打开测试数据文件，生成Workbook对象(wb)
3、根据表单名称选择表单：wb("表单名称")   表单名称为文件左下角名字
4、在表单中，获取单元格数据
    4.1、单元格对象：sh.cell(row,column)  #下标从1开始
    4.2、.value获取单元格的值
    4.3、修改数据：sh.cell(row,column).value = 新值
5、得到当前表单中的总行数和总列数
    sh.max_row  总行数
    sh.mac_colm  总列数
6、在表单当中，获取单元格的数据：
    6.3、修改数据：sh.cell(row,column).value = 新值
7、保存整个工作簿
    WorkBook对象(wb).save("文件路径")
    保存到原文件，需要注意：文件没有被占用
"""

import os
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_cases.xlsx")
print(file_path)

# 1、加载数据文件
from openpyxl import load_workbook
wb = load_workbook(file_path)

# 2、根据表单选择表单：wb["表单名称"]
sh = wb["login"]

# 3、单元格对象：sh.cell(row,column)  #下表从1开始
cel = sh.cell(2,2)
print(cel.value)

# 修改第二行第二列的数据为：liuyue，并且保存到原文件
sh.cell(2,2).value = "liuyue"
wb.save(file_path)

print(sh.max_row)
print(sh.max_column)
